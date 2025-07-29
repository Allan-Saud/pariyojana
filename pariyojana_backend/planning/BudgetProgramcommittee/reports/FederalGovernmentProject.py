# views.py

from rest_framework.views import APIView
from rest_framework.response import Response
from django.db.models import Sum, Count
from planning.BudgetProgramcommittee.FederalGovernmentProject.models import BudgetProgramFederalGovernmentProgram

class BudgetProgramFederalGovernmentProgramChart(APIView):
    def get(self, request):
        # Query all active projects
        projects = BudgetProgramFederalGovernmentProgram.objects.filter(is_deleted=False)

        # 1. thematic_area__name-wise Budget Distribution
        budget_data = (
            projects.values('thematic_area__name')
            .annotate(total_budget=Sum('budget'))
            .order_by('-total_budget')
        )

        total_budget = sum(item['total_budget'] for item in budget_data)

        budget_chart = [
            {
                'label': item['thematic_area__name'],
                'value': float(item['total_budget']),
                'percentage': round((item['total_budget'] / total_budget) * 100, 2) if total_budget > 0 else 0
            }
            for item in budget_data
        ]

        # 2. thematic_area__name-wise Project Count
        count_data = (
            projects.values('thematic_area__name')
            .annotate(project_count=Count('id'))
            .order_by('-project_count')
        )

        total_projects = sum(item['project_count'] for item in count_data)

        count_chart = [
            {
                'label': item['thematic_area__name'],
                'value': item['project_count'],
                'percentage': round((item['project_count'] / total_projects) * 100, 2) if total_projects > 0 else 0
            }
            for item in count_data
        ]

        return Response({
            'budget_distribution': budget_chart,
            'project_count_distribution': count_chart
        })









import io
from django.http import FileResponse, HttpResponse
from django.template.loader import render_to_string
from planning.BudgetProgramcommittee.FederalGovernmentProject.models import BudgetProgramFederalGovernmentProgram
import os
from django.conf import settings
from weasyprint import HTML, CSS
import openpyxl
from openpyxl.styles import Font, Alignment


class BudgetProgramFederalGovernmentProgramDownloadReport(APIView):
    def get_queryset(self, request):
        queryset = BudgetProgramFederalGovernmentProgram.objects.filter(is_deleted=False)
        thematic = request.GET.get('thematic_area')
        sub_thematic = request.GET.get('sub_thematic_area')

        if thematic:
            queryset = queryset.filter(thematic_area_id=thematic)
        if sub_thematic:
            queryset = queryset.filter(sub_area_id=sub_thematic)

        return queryset

    def get(self, request, format=None):
        file_type = request.GET.get('type', 'pdf') 
        queryset = self.get_queryset(request)

        if file_type == 'excel':
            return self.generate_excel(queryset)
        else:
            return self.generate_pdf(request, queryset)

    def generate_excel(self, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"


        ws.merge_cells('A1:F1')
        ws['A1'] = "बर्दगोरिया गाउँपालिका"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Header
        headers = ['क्र.सं.', 'वडा नं', 'योजना', 'क्षेत्र', 'उप-क्षेत्र', 'बजेट']
        ws.append(headers)

        for i, obj in enumerate(queryset, start=1):
            ws.append([
                i,
                obj.ward_no,
                obj.plan_name,
                obj.thematic_area.name if obj.thematic_area else '',
                obj.sub_area.name if obj.sub_area else '',
                obj.budget
            ])

        # Save to stream
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        response = HttpResponse(
            stream,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=federal_government_report.xlsx'
        return response

    

    def generate_pdf(self, request, queryset):
        data = [{
            'index': i + 1,
            'ward_no': obj.ward_no,
            'plan_name': obj.plan_name,
            'thematic_area': obj.thematic_area.name if obj.thematic_area else '',
            'sub_area': obj.sub_area.name if obj.sub_area else '',
            'budget': obj.budget,
        } for i, obj in enumerate(queryset)]
        gov_logo = f'file://{os.path.join(settings.BASE_DIR, "static/images/nepal-govt.png")}'

        html_string = render_to_string('planning/report_template.html', {
            'title': 'बर्दगोरिया गाउँपालिका',
            'projects': data,
            'gov_logo':gov_logo
        })

        css_path = os.path.join(settings.BASE_DIR, 'static', 'css', 'pdf_styles.css')

        # Generate PDF using WeasyPrint
        pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf(
            stylesheets=[CSS(filename=css_path)]
        )

        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="federal_government_report.pdf"'
        return response