# reports/views.py

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from projects.models.project import Project
from django.db.models import Q
from django.utils.encoding import smart_str
import datetime

# Thematic Area
from project_settings.models.thematic_area import ThematicArea
from project_settings.serializers.thematic_area import ThematicAreaSerializer

# Additional Imports
from project_settings.models.sub_thematic_area import SubArea
from project_settings.serializers.sub_thematic_area import SubAreaSerializer

from project_settings.models.source import Source
from project_settings.serializers.source import SourceSerializer

from project_settings.models.expenditure_center import ExpenditureCenter
from project_settings.serializers.expenditure_center import ExpenditureCenterSerializer

from project_settings.models.fiscal_year import FiscalYear
from project_settings.serializers.fiscal_year import FiscalYearSerializer

import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
report_counter = 0

class ReportsDropdownsView(APIView):
    def get(self, request):
        report_types = [
            "सम्पन्न नभएको परियोजनाहरू",
            "परियोजना अनुसार बजेट",
            "प्रगति रिपोर्ट"
        ]
        statuses = [
            "सम्पन्न भएको",
            "सुचारु प्रक्रिया सुनिस्चित भएको",
            "सुरु नभएको"
        ]
        wards = [{"label": f"वडा नं.-{i}", "value": i} for i in range(1, 6)]
        thematic_areas = ThematicAreaSerializer(ThematicArea.objects.filter(is_active=True), many=True).data
        sub_areas = SubAreaSerializer(SubArea.objects.filter(is_active=True), many=True).data
        sources = SourceSerializer(Source.objects.filter(is_active=True), many=True).data
        expenditure_centers = ExpenditureCenterSerializer(ExpenditureCenter.objects.filter(is_active=True), many=True).data
        fiscal_years = FiscalYearSerializer(FiscalYear.objects.all(), many=True).data

        return Response({
            "report_types": report_types,
            "statuses": statuses,
            "wards": wards,
            "thematic_areas": thematic_areas,
            "sub_areas": sub_areas,
            "sources": sources,
            "expenditure_centers": expenditure_centers,
            "fiscal_years": fiscal_years,
        }, status=status.HTTP_200_OK)


class ExportReportExcelView(APIView):
    def post(self, request):
        fiscal_year = request.data.get('fiscal_year')
        status = request.data.get('status')
        ward = request.data.get('ward')
        area = request.data.get('area')
        sub_area = request.data.get('sub_area')
        source = request.data.get('source')
        expenditure_center = request.data.get('expenditure_center')
        report_type = request.data.get('report_type')

        filters = Q(is_active=True, is_deleted=False)

        if fiscal_year:
            filters &= Q(fiscal_year__id=fiscal_year)
        if status:
            filters &= Q(status=status)
        if ward:
            filters &= Q(ward_no__contains=[int(ward)])

        if area:
            filters &= Q(area__id=area)
        if sub_area:
            filters &= Q(sub_area__id=sub_area)
        if source:
            filters &= Q(source__id=source)
        if expenditure_center:
            filters &= Q(expenditure_center__id=expenditure_center)

        if report_type == "सम्पन्न नभएको परियोजनाहरू":
            filters &= ~Q(status="completed")
        elif report_type == "सम्पन्न भएको":
            filters &= Q(status="completed")

        projects = Project.objects.filter(filters).select_related(
            'area', 'sub_area', 'source', 'expenditure_center', 'fiscal_year'
        )

        headers = [
            'ID', 'Project Name', 'Fiscal Year', 'Status', 'Ward No.',
            'Area', 'Sub Area', 'Source', 'Expenditure Center', 'Budget'
        ]

        data = [headers]
       
        for proj in projects:
            try:
                fiscal_year_val = str(proj.fiscal_year.year[0]) if proj.fiscal_year and isinstance(proj.fiscal_year.year, (list, tuple)) else proj.fiscal_year.year if proj.fiscal_year else ''
                status_val = dict(Project.STATUS_CHOICES).get(proj.status, proj.status)
                ward_val = ", ".join([f"वडा नंं. {w}" for w in (proj.ward_no or [])])
                area_val = proj.area.name if proj.area else ''
                sub_area_val = proj.sub_area.name if proj.sub_area else ''
                source_val = proj.source.name if proj.source else ''
                exp_center_val = proj.expenditure_center.name if proj.expenditure_center else ''
                budget_val = float(proj.budget)


                data.append([
                    proj.serial_number,
                    smart_str(proj.project_name),
                    fiscal_year_val,
                    status_val,
                    ward_val,
                    area_val,
                    sub_area_val,
                    source_val,
                    exp_center_val,
                    budget_val,
                ])
            except Exception as e:
                pass



        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Filtered Projects"

        for row in data:
            ws.append(row)

        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = max_length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=projects_report.xlsx'
        now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"projects_report_{now_str}.xlsx"

        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

